import openpyxl
# 导入openpyxl模块后，根据解题步骤，我们首先需要在程序中创建一个新工作簿。
# 使用openpyxl.Workbook()函数即可创建一个新工作簿。
newWB = openpyxl.Workbook()
"""前面的课程我们已经学过，可以访问sheetnames字段来获取工作簿内所有的工作表名称。
通过输出这个字段我们发现，使用openpyxl.Workbook()创建的工作簿里面，都有一张默认的工作表，名称为Sheet。
"""
# print(newWB.sheetnames) 
# ['Sheet']

# 先通过变量newWb使用中括号 + 工作表名称获取这个工作表对象，然后把这个对象赋值给变量aSheet。
aSheet = newWB["Sheet"]
# 接下来，我们将这个工作表的名称修改为“A平台”。
# 通过对工作表对象的.title属性进行赋值，即可修改工作表的名称。
aSheet.title = "A平台"
print(newWB.sheetnames) # ['A平台']

# 通过工作簿对象使用create_sheet()函数可以创建一个名称为Sheet的工作表。
# 若名为Sheet工作表已经存在，则会在Sheet后依次添加数字，比如Sheet1，Sheet2。
# 在创建时如需要指定工作表名称，可以将需要指定的工作表名称作为参数传入create_sheet()函数。
bSheet = newWB.create_sheet("B平台")
cSheet = newWB.create_sheet("C平台")
print(newWB.sheetnames) # ['A平台', 'B平台', 'C平台']

# 在前面Excel读取的课程中，我们已经学习到可以通过“工作表对象["列号行号"].value”这种方式来获取指定的单元格的值。
# 而直接把要设置的值赋值给.value属性，就可以设置或修改这个单元格的值了。
# 为了避免重复的代码，在这里我们可以使用for循环对工作簿对象内的worksheets属性进行遍历，以达到逐个访问所有工作表并设置表头的目的。
for sheet in newWB.worksheets:
    sheet["A1"].value = "编号"
    print(sheet["A1"].value) # 编号

# 通过工作簿对象使用save()函数，将文件保存路径作为参数，即可将工作簿保存到指定的文件路径。 一般地，我们将工作簿存储成后缀名为.xlsx的文件。
newWB.save("/Users/yequ/data/汇总.xlsx")
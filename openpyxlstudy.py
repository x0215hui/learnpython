'''
要使用Python对Excel表格进行读取，我们需要安装一个用于读取数据的工具 openpyxl 。openpyxl 是一个用于读、写Excel文件的开源模块。
安装openpyxl非常简单，在终端中输入代码：pip install openpyxl即可。
如果在自己电脑上安装不上或安装缓慢，可在命令后添加如下配置进行加速：
pip install openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple/
'''
import openpyxl
# 在安装和导入openpyxl之后，读取指定路径的工作簿需要使用函数：openpyxl.load_workbook()。
# 工作簿文件的路径需要作为函数参数传入。
wb = openpyxl.load_workbook("文件路径")

# 得到了工作表名称后，我们可以通过在变量wb后添加中括号[ ]和工作表名称的方式来获得对应的工作表对象。
orderSheet = wb["工作表名称"]
# 如果事先不知道工作簿内有哪些工作表，我们可以通过访问工作簿的 .sheetnames 属性来获取一个包含所有工作表名称的列表。
# 具体操作为在变量wb之后添加代码 .sheetnames 。
# print(wb.sheetnames)

# 要获取工作表中指定的单元格对象，我们可以通过在中括号[ ]内填入列号和行号的方式去获取。
print(orderSheet["A1"]) # <Cell '工作表名称'.A1>
# 输出单元格对象并没有输出单元格内的值。要访问单元格里的值，我们可以在单元格对象后加一个 .value 。
print(orderSheet["A1"].value)
# 若单元格中包含公式，现有方式读取出的值是公式本身。若需要读取公式计算后的值，要在读取工作簿的代码部分，传入一个参数： data_only=True ，便可以得出公式计算后的值了。
wb = openpyxl.load_workbook("文件路径", data_only=True)
print(orderSheet["A1"].value)
# 要对整个工作表的每一行数据进行浏览查询，可以使用for循环对工作表对象的行属性（rows）进行遍历。具体代码为 for rowData in orderSheet.rows:
# 这样程序就会以从上到下的顺序，逐个获取到指定工作表内的每一行数据。
# 可以从运行结果中看到，读取出的每一行数据是由单元格对象组成的元组。
for rowData in orderSheet.rows:
    print(rowData)
    # 行遍历的输出结果，是由单元格对象组成的元组。在元组中要定位到具体的列需要用索引。
    productName = rowData[2].value
    print(productName)

    # 如果要定位的列数字比较大，比如在第I列，通过肉眼观察来确定索引略显繁琐。这时，可以使用函数：openpyxl.utils.cell.column_index_from_string()，来获取列号对应的数字，比如传入参数“E”就会获取到数字5，表示“E”列是第5列。这个数字减一即可得到对应的索引。
    priceIndex = openpyxl.utils.cell.column_index_from_string("I") - 1
    price = rowData[priceIndex].value
    print(price)

"""
对于重复的取表，可以采用定义函数的方法，将上述内容打包为函数。
"""
# 将计算单月销售额的步骤移到函数getMonthlySold中
# 获取单月“火龙果可乐”销售额的函数
# 参数 filePath: 销售数据Excel文件路径
# 返回值: 计算出的销售额结果
def getMonthlySold(filePath):
    # 使用openpyxl.load_workbook()函数读取工作簿
    # 文件路径使用函数参数filePath，然后赋值给变量wb
    # 添加data_only=True打开工作簿，获取公式计算后的值
    wb = openpyxl.load_workbook(filePath, data_only = True)

    # 通过工作簿对象wb获取名为“销售订单数据”的工作表对象，并赋值给变量orderSheet
    orderSheet = wb["销售订单数据"]

    # 定义一个变量colaSold用来表示本月“火龙果可乐”的销售金额
    colaSold = 0

    # 遍历工作表的所有行数据
    for rowData in orderSheet.rows:
        productName = rowData[2].value
        priceIndex = openpyxl.utils.cell.column_index_from_string("I") - 1
        price = rowData[priceIndex].value
        
        if productName == "目标":
            colaSold = colaSold + price

    return colaSold

# 接下来，通过观察销售订单的Excel文件名，我们可以发现，每个文件名仅有月份数字不同。
# 因此，我们可以很方便的使用for循环加range()函数，配合上格式化字符串，来批量生成每个Excel表格的文件路径：2019年{month}月销售订单.xlsx。
# 再把这个文件路径传入到getMonthlySold函数中，来计算各个月份的销售额。最后逐个添加到一个列表soldList中。

soldList = []

for month in range(1,13):
    monthlySold = getMonthlySold(f"2019年{month}月销售订单.xlsx")
    soldList.append(monthlySold)
print(soldList)

# 要获取一个列表中的最大值，可以使用Python内置的max()函数。
maxSold = max(soldList)
print(maxSold)

# 当我们知道了列表中的一个元素，想要去列表中找到这个元素位于什么位置，可以使用列表的index()函数。
# 通过要查询的列表对象使用index()函数，将要查询的元素作为参数传入，则该元素从左往右第一次出现的索引将会被返回。
# 如果查询的元素不在列表中，会报一个ValueError的错误。
maxMonth = soldList.index(maxSold) + 1
print(f"火龙果可乐在{maxMonth}月份卖得最好")